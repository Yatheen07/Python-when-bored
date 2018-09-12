import openpyxl as xl
import smtplib


def getDefaulters(sheets,columnNO):
    unregisteredStudents = {}
    for r in range(2,sheets.max_row + 1):
        status = sheets.cell(row = r,column = columnNO).value
        if status == "unregistered":
            name = sheets.cell(row = r,column = 1).value
            email = sheets.cell(row = r,column = 2).value
            unregisteredStudents[name] = email
            
    return unregisteredStudents
            
sheets = xl.load_workbook(r'./Participants.xlsx')
sheets = sheets['Sheet1']

columnNO = 3

defaulters = getDefaulters(sheets,columnNO)

mail = smtplib.SMTP('smtp.gmail.com',587)
mail.ehlo()

mail.starttls() #tls is transport layer security to keep your login credentials safe.

#login to your mail account
from_email = "yourMailID"
password = "yourPassWord!"
mail.login(from_email,password)


for name , email in defaulters.items():
    content = "Group Mail test! Hello Mr/Mrs %s" % (name)
    mail_body = """From: """+from_email+"""\nTo: """+email+"""\nSubject: Form Registration Incomplete!\n"""+content
    print("Sending mail to %s.." % email)
    status = mail.sendmail(from_email,email,mail_body)
    if status != {}:
        print("There was a problem sending mail to %s: %s" %(email,status))
    
mail.quit()
